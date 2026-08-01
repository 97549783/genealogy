"""Проверки семантической структуры одной научной школы."""

import numpy as np
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

from core.semantic.models import build_section_selection
from tabs.school_analysis.semantic import (
    build_school_semantic_dataset, compute_branch_semantics,
    compute_generation_semantics, compute_school_heterogeneity,
    compute_school_semantic_center,
)
from tabs.school_analysis.exports import build_excel_report


def _dataset(count: int = 6):
    angles = np.linspace(0, 0.9, count)
    matrix = np.array([[np.cos(value), np.sin(value)] for value in angles], dtype=np.float32)
    codes = [str(index) for index in range(count)]
    section_index = pd.DataFrame({
        "Code": codes, "section_key": ["research_goal"] * count,
        "matrix_row": range(count), "text_id": [f"t{code}" for code in codes],
    })
    metadata = pd.DataFrame({
        "Code": codes, "candidate_name": [f"Автор {code}" for code in codes],
        "title": [f"Работа {code}" for code in codes], "year": range(2000, 2000 + count),
        "Поколение": [1] * 3 + [2] * (count - 3),
    })
    selection = build_section_selection("selected", ["research_goal"], min_coverage=1)
    dataset = build_school_semantic_dataset(
        root="Школа", member_codes=codes, section_index=section_index, matrix=matrix,
        selection=selection, normalized=True, dissertation_metadata=metadata,
    )
    return dataset, selection


def test_center_heterogeneity_medoid_and_categories() -> None:
    dataset, selection = _dataset(6)
    center = compute_school_semantic_center(dataset, selection)
    assert np.isclose(np.linalg.norm(center["research_goal"]), 1)
    result = compute_school_heterogeneity(dataset, center, selection)
    assert result.medoid_code in {"2", "3"}
    assert result.summary["Показатель"].str.contains("90-й процентиль").any()
    assert result.dissertation_distances["Категория"].notna().all()


def test_fewer_than_five_has_no_categories() -> None:
    dataset, selection = _dataset(4)
    result = compute_school_heterogeneity(dataset, compute_school_semantic_center(dataset, selection), selection)
    assert result.dissertation_distances["Категория"].isna().all()


def test_generation_small_sample_rules_and_continuity() -> None:
    dataset, selection = _dataset(6)
    result = compute_generation_semantics(dataset, {1: {"0", "1"}, 2: {"2", "3", "4", "5"}}, selection)
    first, second = result.summary.iloc[0], result.summary.iloc[1]
    assert pd.isna(first["Тематическая неоднородность"])
    assert second["Сходство с предыдущим поколением"] is not None
    assert any("недостаточно" in value for value in result.diagnostics)


def test_branch_overlap_is_excluded_only_from_silhouette() -> None:
    dataset, selection = _dataset(8)
    branches = {"Ветвь А": {"0", "1", "2", "3", "7"}, "Ветвь Б": {"4", "5", "6", "7"}}
    result = compute_branch_semantics(dataset, branches, selection)
    assert result.summary.set_index("Ветвь").loc["Ветвь А", "Диссертаций с векторами"] == 5
    assert result.ambiguous_dissertations["Code"].tolist() == ["7"]
    assert result.silhouette_overall is not None
    assert set(result.similarity_matrix.columns) >= {"Ветвь", "Ветвь А", "Ветвь Б"}


def test_optional_semantic_excel_sheets_preserve_existing_report() -> None:
    empty = pd.DataFrame()
    payload = build_excel_report(
        metrics_df=pd.DataFrame({"Показатель": ["Всего"], "Значение": [1]}),
        generations_df=empty, yearly_df=empty, city_df=empty, institutional={},
        opponents_df=empty, continuity_df=empty,
        semantic_summary=pd.DataFrame({"Показатель": ["Медиана"], "Значение": [.2]}),
        semantic_branches=pd.DataFrame({"Ветвь": ["А"]}),
    )
    sheets = load_workbook(BytesIO(payload), read_only=True).sheetnames
    assert sheets == ["Метрики", "Семантическая сводка", "Семантика ветвей"]


def test_zero_eligible_dissertations_returns_stable_schema() -> None:
    selection = build_section_selection("selected", ["research_goal"], min_coverage=1)
    dataset = build_school_semantic_dataset(
        root="Пустая школа", member_codes={"1"}, section_index=pd.DataFrame({
            "Code": [], "section_key": [], "matrix_row": [],
        }), matrix=np.array([[1, 0]], dtype=np.float32), selection=selection,
        normalized=True, dissertation_metadata=pd.DataFrame({"Code": ["1"]}),
    )
    result = compute_school_heterogeneity(dataset, {}, selection)
    assert result.medoid_code is None
    assert result.dissertation_distances.columns.tolist()[-1] == "Code"
    assert result.dissertation_distances.empty
    assert "Нет диссертаций" in result.diagnostics[0]
    assert dataset.metadata.iloc[0]["Code"] == "1"


def test_excluded_group_member_keeps_metadata() -> None:
    selection = build_section_selection("selected", ["research_goal", "research_methods"], min_coverage=1)
    dataset = build_school_semantic_dataset(
        root="Школа", member_codes={"1"},
        section_index=pd.DataFrame({"Code": ["1"], "section_key": ["research_goal"], "matrix_row": [0]}),
        matrix=np.array([[1, 0]], dtype=np.float32), selection=selection, normalized=True,
        dissertation_metadata=pd.DataFrame({"Code": ["1"], "candidate_name": ["Автор"], "title": ["Название"], "year": [2020], "Поколение": [2]}),
    )
    generations = compute_generation_semantics(dataset, {2: {"1"}}, selection)
    assert generations.dissertation_details[2].iloc[0]["candidate_name"] == "Автор"
    assert generations.dissertation_details[2].iloc[0]["Причина исключения"] == "Недостаточное покрытие выбранных разделов"


def test_semantic_export_does_not_multiply_existing_percentage_twice() -> None:
    empty = pd.DataFrame()
    payload = build_excel_report(
        metrics_df=pd.DataFrame({"Показатель": ["Всего"], "Значение": [1]}),
        generations_df=empty, yearly_df=empty, city_df=empty, institutional={},
        opponents_df=empty, continuity_df=empty,
        semantic_generation_dissertations=pd.DataFrame({"Автор": ["А"], "Покрытие, %": [60.0], "Code": ["1"]}),
    )
    sheet = load_workbook(BytesIO(payload), data_only=True)["Диссертации поколений"]
    assert sheet.cell(2, 2).value == 60.0
    assert "Code" not in [cell.value for cell in sheet[1]]


def test_semantically_excluded_ambiguous_member_remains_visible() -> None:
    selection = build_section_selection("selected", ["research_goal", "research_methods"], min_coverage=1)
    dataset = build_school_semantic_dataset(
        root="Школа", member_codes={"1", "2", "3", "4", "5", "6", "X"},
        section_index=pd.DataFrame({
            "Code": list("123456"), "section_key": ["research_goal"] * 6,
            "matrix_row": range(6),
        }), matrix=np.eye(6, dtype=np.float32), selection=selection, normalized=True,
        dissertation_metadata=pd.DataFrame({"Code": [*list("123456"), "X"], "candidate_name": [*list("АБВГДЕ"), "Ж"]}),
    )
    result = compute_branch_semantics(dataset, {"А": {"1", "2", "3", "X"}, "Б": {"4", "5", "6", "X"}}, selection)
    ambiguous = result.ambiguous_dissertations.set_index("Code")
    assert "X" in ambiguous.index
    assert ambiguous.loc["X", "Ветви"] == "А; Б"
    assert "недостаточное покрытие" in ambiguous.loc["X", "Причина исключения"]


def test_invalid_ambiguous_member_has_actual_exclusion_reason() -> None:
    selection = build_section_selection("selected", ["research_goal"], min_coverage=1)
    dataset = build_school_semantic_dataset(
        root="Школа", member_codes={"1", "2", "X"},
        section_index=pd.DataFrame({
            "Code": ["1", "2", "X"], "section_key": ["research_goal"] * 3,
            "matrix_row": [0, 1, 99],
        }), matrix=np.eye(2, dtype=np.float32), selection=selection, normalized=True,
        dissertation_metadata=pd.DataFrame({"Code": ["1", "2", "X"], "candidate_name": ["А", "Б", "В"]}),
    )
    result = compute_branch_semantics(dataset, {"А": {"1", "X"}, "Б": {"2", "X"}}, selection)
    reason = result.ambiguous_dissertations.set_index("Code").loc["X", "Причина исключения"]
    assert "недопустимые или нулевые строки" in reason
    assert "исключено только" not in reason
