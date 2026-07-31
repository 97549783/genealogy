"""Проверки сравнения школ по характеристикам диссертаций."""

from io import BytesIO

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from core.semantic.models import build_section_selection
from tabs.school_comparison.exports import build_semantic_school_comparison_excel
from tabs.school_comparison.semantic import (
    compute_per_section_silhouette, compute_semantic_school_comparison,
    gather_semantic_school_dataset,
)


def _dataset(root: str, codes: list[str], rows: list[int], matrix: np.ndarray):
    selection = build_section_selection("selected", ["research_goal"], min_coverage=1)
    index = pd.DataFrame({
        "Code": codes, "section_key": ["research_goal"] * len(codes),
        "matrix_row": rows, "text_id": [f"t{row}" for row in rows],
    })
    metadata = pd.DataFrame({"Code": codes, "candidate_name": [f"Автор {code}" for code in codes]})
    return gather_semantic_school_dataset(
        root=root, member_codes=codes, section_index=index, matrix=matrix,
        selection=selection, normalized=True, metadata_df=metadata,
    )


def test_overlapping_code_is_preserved_under_each_school() -> None:
    matrix = np.array([[1, 0], [.9, .1], [0, 1], [.1, .9]], dtype=np.float32)
    datasets = {
        "А": _dataset("А", ["общая", "а"], [0, 1], matrix),
        "Б": _dataset("Б", ["общая", "б"], [2, 3], matrix),
    }
    selection = build_section_selection("selected", ["research_goal"], min_coverage=1)
    result = compute_semantic_school_comparison(datasets=datasets, selection=selection, distance_batch_size=2)
    assert result.overall_silhouette is not None
    assert result.dissertation_silhouettes["Code"].tolist().count("общая") == 2
    assert set(result.school_summary["Научная школа"]) == {"А", "Б"}


def test_coverage_exclusion_and_undefined_pairs() -> None:
    matrix = np.array([[1, 0], [0, 1]], dtype=np.float32)
    selection = build_section_selection("selected", ["research_goal", "research_methods"], min_coverage=.6)
    index = pd.DataFrame({"Code": ["1"], "section_key": ["research_goal"], "matrix_row": [0]})
    dataset = gather_semantic_school_dataset(
        root="А", member_codes={"1"}, section_index=index, matrix=matrix,
        selection=selection, normalized=True, metadata_df=pd.DataFrame({"Code": ["1"]}),
    )
    assert not dataset.dissertation_vectors
    assert dataset.excluded.iloc[0]["Причина исключения"].startswith("Недостаточное")


def test_per_section_eligibility_and_excel_labels() -> None:
    matrix = np.array([[1, 0], [.9, .1], [0, 1], [.1, .9]], dtype=np.float32)
    datasets = {"А": _dataset("А", ["1", "2"], [0, 1], matrix), "Б": _dataset("Б", ["3", "4"], [2, 3], matrix)}
    section_selection = build_section_selection("selected", ["research_goal", "research_methods"], min_coverage=.5)
    section = compute_per_section_silhouette(datasets=datasets, selection=section_selection)
    assert section.columns.tolist() == ["Раздел характеристики", "Коэффициент силуэта", "Число школ", "Число диссертаций", "Полнота данных, %", "Статус"]
    selection = build_section_selection("selected", ["research_goal"], min_coverage=1)
    result = compute_semantic_school_comparison(datasets=datasets, selection=selection, distance_batch_size=2)
    workbook = load_workbook(BytesIO(build_semantic_school_comparison_excel(result, {
        "representation": "characteristics", "sections_mode": "all", "minimum_coverage": .6,
    })), read_only=True)
    assert workbook.sheetnames == ["Параметры", "Сводка по школам", "Силуэт по диссертациям", "Силуэт по разделам", "Исключённые диссертации", "Диагностика"]
    parameters = list(workbook["Параметры"].values)
    assert any("Все доступные характеристики" in str(row) for row in parameters)
    assert any("60.0 %" in str(row) for row in parameters)
    diagnostics = " ".join(str(row) for row in workbook["Диагностика"].values)
    assert "Причина попарного расчёта: Расчёт выполнен" in diagnostics


def test_partial_export_does_not_claim_pairwise_success() -> None:
    matrix = np.array([[1, 0]], dtype=np.float32)
    datasets = {"А": _dataset("А", ["1"], [0], matrix)}
    selection = build_section_selection("selected", ["research_goal"])
    result = compute_semantic_school_comparison(datasets=datasets, selection=selection, distance_batch_size=2)
    assert result.pairwise_diagnostics.reason == "insufficient_samples"
    workbook = load_workbook(BytesIO(build_semantic_school_comparison_excel(result, {})), read_only=True)
    diagnostics = " ".join(str(row) for row in workbook["Диагностика"].values)
    assert "Недостаточно данных для попарного расчёта" in diagnostics
